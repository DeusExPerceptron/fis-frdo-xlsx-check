import openpyxl
import os
import datetime as dt
import frdo_data as fd
from typing import List, Any, Optional


class CheckResult:
    def __init__(self, result: bool, msg: Any):
        self.result = result
        self.msg = msg


def is_correct_snils(value: str) -> CheckResult:  # XXX-XXX-XXX XX
    result: bool = True
    result_msg: List[str] = []

    if ((not isinstance(value, str) or len(str(value)) != 14)
       or value[3] + value[7] + value[11] != '-- '
       or not (value[0:3] + value[4:7] + value[8:11] + value[12:]).isdigit()):
        return CheckResult(False, 'неверный формат СНИЛС (XXX-XXX-XXX XX)')
    else:
        snils = value[0:3] + value[4:7] + value[8:11] + value[12:]
        snils_sum = sum(map(lambda s_x, s_mult: int(s_x) * s_mult, snils[-3::-1], range(1, 10)))

    while snils_sum > 101:
        snils_sum %= 101

    if snils_sum in (100, 101):
        snils_sum = 0

    if snils_sum != int(snils[-2:]):
        return CheckResult(False, f'ошибка контрольной суммы СНИЛС ({snils_sum} != {snils[-2:]})')

    return CheckResult(result, result_msg)


def is_correct_char_in_name(value: str) -> bool:
    rus_lower_chr = [chr(i) for i in range(ord('а'), ord('я') + 1)] + ['ё']
    for i in str(value):
        if i.lower() not in rus_lower_chr + list(" -"):
            return False
    return True


def is_empty_value(value: Optional[str]) -> bool:
    if value is None or (isinstance(value, str) and (value == '')):
        return True
    return False


def is_correct_value(value: Optional[int | str | dt.datetime], xlsx_check: Optional[str] = None,
                     can_be_empty: bool = False, value_type: Optional[str] = None,
                     min_v: Optional[int | dt.date] = None, max_v: Optional[int | dt.date] = None) -> CheckResult:
    result: bool = True
    result_msg: List[str] = []

    if not can_be_empty and is_empty_value(value):
        return CheckResult(False, ['пустое значение'])

    if isinstance(value, int):
        value = str(value)

    if isinstance(value, str) and len(value) > 0:
        if value.startswith(' '):
            result = False
            result_msg.append('содержит пробел в начале')
        if value.endswith(' '):
            result = False
            result_msg.append('содержит пробел в конце')
        if '\n' in value:
            result = False
            result_msg.append('содержит перенос на новую строку')

        if result and value_type is not None:
            if value_type == 'alpha' and not is_correct_char_in_name(value):
                result = False
                result_msg.append('значение содержит символы кроме кириллицы, - или пробела')
            elif value_type == 'digit':
                if not value.isdigit():
                    result = False
                    result_msg.append('значение не число')
                else:
                    if isinstance(min_v, int) and int(value) < min_v:
                        result = False
                        result_msg.append(f'число меньше минимального значения {min_v}')
                    if isinstance(max_v, int) and int(value) > max_v:
                        result = False
                        result_msg.append(f'число больше максимального значения {max_v}')

        if value_type == 'SNILS':
            snils_check = is_correct_snils(value.strip().replace('\n', ''))
            result = snils_check.result
            result_msg.append(snils_check.msg)

    if value_type == 'date':
        if not isinstance(value, dt.datetime):
            result = False
            result_msg.append('неверный формат даты (текстовый формат ячейки)')
        else:
            if isinstance(min_v, dt.date) and value.date() < min_v:
                result = False
                result_msg.append(f'дата меньше минимального значения {min_v.strftime("%d.%m.%Y")}')
            if isinstance(max_v, dt.date) and value.date() > max_v:
                result = False
                result_msg.append(f'дата больше максимального значения {max_v.strftime("%d.%m.%Y")}')

    if xlsx_check is not None:
        if str(value) not in fd.xlsx_check_dict[xlsx_check]:
            result = False
            result_msg.append(f'нет в разрешенных значениях шаблона')

    return CheckResult(result, result_msg)


def add_log_msg(fn_log_file: str, msg: str):
    with open(fn_log_file, 'a', encoding='utf-8') as log_file:
        log_file.write(f'{msg}\n')


fn_all_files = os.listdir()
fn_xlsx_list = list(filter(lambda x: (x.endswith('.xlsx') and not x.startswith('~')), fn_all_files))

for xlsx_file in fn_xlsx_list:
    rows_with_data_count = 0
    empty_row_found = False
    empty_row_before_end = False

    print(f'***{xlsx_file}***')

    fn_log = f'{xlsx_file.rstrip(".xlsx")}.log'

    with open(fn_log, 'w', encoding='utf-8') as log_file_t:
        pass

    try:
        xlsx_wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    except Exception as e:
        add_log_msg(fn_log, f'не удалось открыть файл: {xlsx_file}')
        add_log_msg(fn_log, str(e))
        continue

    wb_xlsx_sheet_name = 'Шаблон'

    if wb_xlsx_sheet_name not in xlsx_wb.sheetnames:
        add_log_msg(fn_log, f'не удалось найти вкладку: {wb_xlsx_sheet_name}')
        continue

    try:
        xlsx_ws = xlsx_wb[wb_xlsx_sheet_name]
    except Exception as e:
        add_log_msg(fn_log, f'не удалось открыть вкладку: {wb_xlsx_sheet_name}')
        add_log_msg(fn_log, str(e))
        continue

    for col_i, col_name in fd.cols_name.items():
        if xlsx_ws[1][col_i].value.strip() != col_name:
            add_log_msg(fn_log, 'нарушен порядок столбцов или их заголовки')
            continue

    for i_row, row in enumerate(xlsx_ws.iter_rows(min_row=2)):
        errors_dict = {}

        if is_empty_value(row[0].value):
            empty_row_found = True
            continue
        if empty_row_found:
            empty_row_before_end = True

        for chk_i in fd.cols_name.keys():
            i_xlsx_check: Optional[str] = None
            i_can_be_empty: bool = False
            i_value_type: Optional[str] = None
            i_min_v: Optional[int | dt.date] = None
            i_max_v: Optional[int | dt.date] = None

            if chk_i in (1, 2, 3, 4, 5, 6, 19, 20, 21, 22, 23):
                i_xlsx_check = fd.cols_name[chk_i]
            if chk_i in (16, 24, 25, 26, 27, 28, 29, 30, 31, 32):
                i_can_be_empty = True
            if chk_i in (14, 15, 16):
                i_value_type = 'alpha'
            if chk_i in (11, 12, 13):
                i_value_type = 'digit'
                if chk_i in (11, 12):
                    i_min_v, i_max_v = 1900, 2100
                elif chk_i == 13:
                    i_min_v, i_max_v = 0, 15

            if chk_i == 18:
                i_value_type = 'SNILS'

            if chk_i in (9, 17):
                i_value_type = 'date'
                if chk_i == 9:
                    i_min_v, i_max_v = dt.date(1950, 1, 1), dt.date(2100, 12, 31)
                elif chk_i == 17:
                    i_min_v, i_max_v = dt.date(1900, 1, 1), dt.date(2100, 12, 31)

            chk_status = is_correct_value(row[chk_i].value, xlsx_check=i_xlsx_check, can_be_empty=i_can_be_empty,
                                          value_type=i_value_type, min_v=i_min_v, max_v=i_max_v)
            if not chk_status.result:
                errors_dict.setdefault(chk_i, []).extend(chk_status.msg)

        if len(errors_dict) > 0:
            add_log_msg(fn_log, f'Строка {i_row + 2}:')
        for error_i, error_msg in errors_dict.items():
            t_value = repr(row[error_i].value)
            if isinstance(row[error_i].value, dt.datetime):
                t_value = row[error_i].value.strftime("%d.%m.%Y")
            add_log_msg(fn_log, f'\tСтолбец {(error_i + 1):02}: {fd.cols_name[error_i]} | '
                                f'{t_value} | {", ".join(error_msg)}')

        rows_with_data_count += 1

    add_log_msg(fn_log, '')

    if empty_row_before_end:
        add_log_msg(fn_log, 'в файле есть пустые строки между строками с данными')

    add_log_msg(fn_log, f'записей проверено (за исключением пустых): {rows_with_data_count}')

    xlsx_ws = None
    xlsx_wb = None
